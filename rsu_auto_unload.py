import datetime
import logging
import os
import time
from selenium.webdriver.support import expected_conditions
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.by import By
import warnings
from selenium.webdriver.support.wait import WebDriverWait
import json


def auto_unload():
    warnings.filterwarnings('ignore')  # игнорирование предупреждений
    # logging # TODO добавить логи вместо prints
    logging.basicConfig(level=logging.ERROR, filename='log.txt', filemode="a+",
                        format="%(asctime)-15s %(levelname)-8s %(message)s")
    logger = logging.getLogger('logger')

    # инициализация словаря предыдущих значений
    rsu_previews_data = {'1-1': set(), '2-1': set(), '3-1': set(), '4-1': set(), '4-2': set(), '4-3': set(),
                         '4-4': set(), '5-1': set(), '6-1': set()}
    # словарь РСУ
    rsu = {'1-1': {'ip': 'https://192.168.10.111/documentation/documentation.html', 'location': 'R1'},
           '2-1': {'ip': 'https://192.168.10.138/documentation/documentation.html', 'location': 'R1'},
           '3-1': {'ip': 'https://192.168.10.113/documentation/documentation.html', 'location': 'R2'},
           '4-1': {'ip': 'https://192.168.10.114/documentation/documentation.html', 'location': 'R2'},
           '4-2': {'ip': 'https://192.168.10.115/documentation/documentation.html', 'location': 'R2'},
           '4-3': {'ip': 'https://192.168.10.116/documentation/documentation.html', 'location': 'R2'},
           '4-4': {'ip': 'https://192.168.10.117/documentation/documentation.html', 'location': 'R2'},
           '5-1': {'ip': 'https://192.168.10.137/documentation/documentation.html', 'location': 'R1'},
           '6-1': {'ip': 'https://192.168.10.112/documentation/documentation.html', 'location': 'R2'},
           }

    # бесконечный цикл для опроса РСУ
    while True:
        time_now_str = datetime.datetime.now().strftime("%H:%M")
        time_now = datetime.datetime.strptime(time_now_str, "%H:%M")
        today = datetime.datetime.now().strftime('%d.%m.%Y')  # сегодня
        print(f'Запуск в {time_now_str}. {today}.')
        PID = os.getpid()  # id для определения процессов
        PARENT_PID = os.getppid()  # id для определения родительских процессов
        print(f'PID = {PID}')
        print(f'Parent PID = {PARENT_PID}')
        feedback_list = {'datetime': str(datetime.datetime.now()), 'PID': PID, 'PARENT_PID': PARENT_PID}
        try:  # создание файла обратной связи для контроля зависания
            r_filename = r"O:\Расчет эффективности\Выгрузки РСУ\r.json"  # файл контроля зависания
            with open(r_filename, 'w') as file:
                file.write(json.dumps(feedback_list))
                print('Файл обратной связи создан.\n')
                # file.write(str(feedback_list))
                # file.write(str(datetime.datetime.now()))
                # file.write(str(PID))
                # file.write(str(PARENT_PID))
        except Exception as e:
            print(e, 'Ошибка создания файла обратной связи!\n')
        try:  # закрытие драйвера
            driver.close()
        except Exception as e:
            print(e, 'Драйвер уже закрыт.')
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")  # включение headless режима
        options.add_argument('--ignore-certificate-errors')  # отключение сообщений ошибки сертификатов
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        try:
            driver = webdriver.Chrome(options=options)  # назначение драйвера
        except Exception as e:
            print(e, 'Ошибка создания webdriver. Отсутствует связь.\n')
        try:
            driver.implicitly_wait(10)
        except Exception as e:
            print(e, 'Ошибка timeout!')
        for rsu_name in rsu.keys():
            # result_file_name = rf"{os.getcwd()}\Выгрузка РСУ {rsu_name} за {today}.xlsx"  # имя xlsx файла
            result_file_name = rf"O:\Расчет эффективности\Выгрузки РСУ\Выгрузка РСУ {rsu_name} за {today}.xlsx"
            rsu_result_list = []  # список результатов
            time_sum = 0  # сумма дуги
            try:
                print('Запуск для:', rsu[rsu_name]['ip'], rsu_name, time_now_str, today)
                try:
                    req = requests.get(rsu[rsu_name]['ip'], verify=False)  # проверка включенности аппарата
                    req_status = True
                except Exception as e:
                    print(e)
                    req = None
                    req_status = False
                    print(rsu_name, 'Недоступен.\n')
                if req_status:  # если страница открылась
                    is_updated = False  # переменная факта обновления данных
                    if req.status_code == 200:  # если страница загрузилась
                        # если файл есть, то создаем, иначе открываем
                        if os.path.isfile(result_file_name):
                            res_wb = openpyxl.load_workbook(result_file_name, data_only=True)
                            res_sh = res_wb['Выгрузка РСУ']
                            sum_result_sh = res_wb['Сумма дуги РСУ']
                            # Сохранение существующего файла в словарь предыдущих данных
                            arc_sum_before = sum_result_sh['H2'].value  # предыдущее значение дуги
                            print(arc_sum_before)
                            for row in res_sh.iter_rows(min_row=2, min_col=1, max_row=res_sh.max_row,
                                                        max_col=res_sh.max_column, values_only=True):
                                rsu_previews_data[rsu_name].add(row[5])
                                # print(row[5])
                            if rsu_previews_data:
                                print('Существующие данные перенесены.\n')
                            else:
                                print('Данные для переноса отсутствуют.\n')
                        else:
                            res_wb = openpyxl.Workbook()
                            res_sh = res_wb.create_sheet('Выгрузка РСУ')  # листа выгрузки
                            sum_result_sh = res_wb.create_sheet('Сумма дуги РСУ')  # лист суммы дуги
                            res_wb.remove(res_wb['Sheet'])  # удаление лишнего листа
                            # Шапка листа выгрузки
                            res_sh.append(
                                ['имя РСУ', 'Дата', 'Время операции', 'Время дуги, с', 'Полная строка выгрузки'])
                            # Шапка листа суммы
                            sum_result_sh.append(['Дата', '', '', '', '', '', '', 'время дуги',
                                                  'стоимость минуты дуги'])
                            arc_sum_before = 0
                        driver.get(rsu[rsu_name]['ip'])
                        # time.sleep(30)
                        # ожидание появления элемента
                        element = WebDriverWait(driver, 30).until(
                            expected_conditions.presence_of_element_located((By.ID, "tablecontainer")))
                        driver.implicitly_wait(20)  # неявное ожидание 20 сек
                        data = driver.find_elements(By.CLASS_NAME, 'rowElement')
                        for row in data:  # обработка строк для отчёта
                            line = row.text  # текст элемента
                            if line not in rsu_previews_data[rsu_name]:
                                # print(line)
                                coma_index = line.find(',')  # индекс запятой
                                if coma_index != -1:
                                    arc_date = line[coma_index - 10:coma_index]
                                else:
                                    arc_date = 0
                                # индекс строки до первого разрыва страницы-отделение интервала времени работы аппарата
                                break_index = line.find('\n')
                                if break_index != -1:
                                    arc_period = line[coma_index + 2:break_index]
                                else:
                                    arc_period = 0
                                # индекс до второго разрыва страницы - отделение времени горения дуги
                                break_index_2 = line.find('\n', break_index + 1, len(line))
                                if break_index_2 != -1:
                                    arc_time = line[break_index + 1:break_index_2 - 2]
                                else:
                                    arc_time = 0
                                if str(arc_date).strip() == str(today).strip():
                                    rsu_result_list.append([rsu_name, arc_date, arc_period, float(arc_time), '',
                                                            line])
                                    # line.replace('\n', '')])
                                    time_sum = time_sum + float(arc_time)
                                    res_sh.append([rsu_name, arc_date, arc_period, float(arc_time), '',
                                                   line])  # строка в исходном виде для сравнения
                                    # line.replace('\n', '')]) #
                                    rsu_previews_data[rsu_name].add(line)  # добавление строки в уже выгруженные
                                    print(f'Заполнено.')
                                    # print(f'Заполнено: {line}')
                                    is_updated = True
                            else:
                                print(f'Данные для {rsu_name} в не изменились.')
                        # print(rsu_result_list)
                        # print(rsu_previews_data[rsu_name])
                        arc_value = time_sum / 60 / 60 + float(arc_sum_before)  # значение дуги
                        print(arc_value, rsu_name)
                        # заполнение листа результатов
                        sum_result_sh['A2'] = today
                        sum_result_sh['D2'] = rsu_name
                        sum_result_sh['G2'] = rsu[rsu_name]['location']
                        # print(rsu[rsu_name]['location'])
                        sum_result_sh['H2'] = round(arc_value, 2)
                        sum_result_sh['I2'] = 4
                        # строка со значениями итераций дял тестов
                        sum_result_sh['J2'] = str(sum_result_sh['J2'].value) + '+' + f"{round(time_sum / 60 / 60, 2)}"
                        time.sleep(5)
                        if is_updated:
                            res_wb.save(result_file_name)
                            print(f'Файл {result_file_name} обновлен.\n')
                        else:
                            print(f'Данные не изменились. Файл {result_file_name} НЕ обновлен.\n')
                        ActionBuilder(driver).clear_actions()  # сброс значений действий драйвера
                    else:
                        print(rsu_name, 'Недоступен.')

            except Exception as e:
                print(e)
        reset_time_interval_start = datetime.datetime.strptime("00:01", "%H:%M")
        reset_time_interval_end = datetime.datetime.strptime("00:21", "%H:%M")
        # сброс previews data
        if reset_time_interval_end > time_now > reset_time_interval_start:
            try:
                rsu_previews_data = {'1-1': set(), '2-1': set(), '3-1': set(), '4-1': set(), '4-2': set(),
                                     '4-3': set(),
                                     '4-4': set(), '5-1': set(), '6-1': set()}
                print(f'rsu_previews_data reset complete!')
            except Exception as e:
                print(e, "\nrsu_previews_data reset NOT complete!\n")
        print(f'Итерация пройдена {time_now_str} {today}, ожидание следующей. Запуск через 2 минуты.\n')
        time.sleep(2 * 60)  # период ожидания след итерации


if __name__ == '__main__':
    auto_unload()
